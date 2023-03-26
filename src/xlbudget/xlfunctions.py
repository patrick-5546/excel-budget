import calendar
from logging import getLogger

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = getLogger(__name__)

FORMAT_ACCOUNTING = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FORMAT_DATE = "MM/DD/YYYY"

MONTH_NAME_0_IND = calendar.month_name[1:]


def create_year_sheet(wb: Workbook, year: str):
    """Creates a year sheet, with a table for each month.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The workbook to create the sheet in.
        year (str): The year.
    """
    ws = wb.create_sheet(year)
    num_tables = len(MONTH_NAME_0_IND)

    table_headers = ["Date", "Description", "Amount"]
    table_formats = [FORMAT_DATE, None, FORMAT_ACCOUNTING]
    for c_start in range(
        1, (len(table_headers) + 1) * num_tables + 1, len(table_headers) + 1
    ):
        month_ind = c_start // (len(table_headers) + 1)
        month = MONTH_NAME_0_IND[month_ind]
        logger.debug(f"creating {month} table")

        # table title
        ws.cell(row=1, column=c_start, value=month)
        ws.merge_cells(
            start_row=1,
            start_column=c_start,
            end_row=1,
            end_column=c_start + len(table_headers) - 2,
        )

        # table sum
        sum = ws.cell(
            row=1,
            column=c_start + len(table_headers) - 1,
            value=f"=SUM({month}[{table_headers[-1]}])",
        )
        sum.number_format = FORMAT_ACCOUNTING
        logger.debug(f"created sum cell {sum.coordinate}='{sum.value}'")

        # table header and formating
        for i in range(len(table_headers)):
            c = c_start + i

            # header
            ws.cell(row=2, column=c, value=table_headers[i])

            # formatting
            cell = ws.cell(row=3, column=c)
            if table_formats[i]:
                cell.number_format = table_formats[i]

        # create table
        c_start_ltr = get_column_letter(c_start)
        c_end_ltr = get_column_letter(c_start + len(table_headers) - 1)
        ref = f"{c_start_ltr}2:{c_end_ltr}3"
        logger.debug(f"table {ref=}")
        tab = Table(displayName=month, ref=ref)

        # add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style

        ws.add_table(tab)
