"""The commands, implemented as implementations of the abstract class `Command`."""

import os
import sys
from abc import ABC, abstractmethod
from argparse import ArgumentParser, Namespace, _SubParsersAction
from logging import getLogger
from typing import List, Optional, Type

from openpyxl import Workbook, load_workbook

from xlbudget.inputformat import GetInputFormats, InputFormat, parse_input
from xlbudget.rwxlb import update_xlbudget

logger = getLogger(__name__)


class Command(ABC):
    """The abstract class that the command implementations implement.

    Attributes: Class Attributes
        default_path (str): The default path of the xlbudget file.

    Attributes:
        trial (bool): If True, the xlbudget file will not be written to.
        path (str): The path to the xlbudget file.
    """

    default_path: str = "xlbudget.xlsx"

    @property
    @abstractmethod
    def name(self) -> str:
        """Ensures that the `name` class attribute is defined in subclasses.
        Part 1/2 of the abstract attribute implementation of `name`.
        Reference: https://stackoverflow.com/a/53417582.
        """
        raise NotImplementedError

    def get_name(self) -> str:
        """Used to access the `name` class attribute defined in subclasses.
        Part 2/2 of the abstract attribute implementation of `name`.
        Reference: https://stackoverflow.com/a/53417582.
        """
        return self.name

    @property
    @abstractmethod
    def aliases(self) -> List[str]:
        """Ensures that the `aliases` class attribute is defined in subclasses.
        Part 1/2 of the abstract attribute implementation of `aliases`.
        Reference: https://stackoverflow.com/a/53417582.
        """
        raise NotImplementedError

    def get_aliases(self) -> List[str]:
        """Used to access the `aliases` class attribute defined in subclasses.
        Part 2/2 of the abstract attribute implementation of `aliases`.
        Reference: https://stackoverflow.com/a/53417582.
        """
        return self.aliases

    @classmethod
    def configure_common_args(cls, parser: ArgumentParser) -> None:
        """Configures the arguments that are used by all commands.

        Args:
            parser (ArgumentParser): The argument parser.
        """
        parser.add_argument(
            "-t",
            "--trial",
            action="store_true",
            help="try a command without generating/updating the xlbudget file",
        )
        parser.add_argument(
            "-p",
            "--path",
            help="path to the xlbudget file (default: %(default)s)",
            default=cls.default_path,
        )

    @classmethod
    @abstractmethod
    def configure_args(cls, subparsers: _SubParsersAction) -> None:
        pass

    @abstractmethod
    def __init__(self, args: Namespace) -> None:
        self.trial = args.trial

        self._check_path(args.path)
        self.path = args.path

    @staticmethod
    def _check_path(path: str) -> None:
        """Check that `path` is a valid path to an xlbudget file.

        Args:
            path (str): The xlbudget path.

        Raises:
            ValueError: If `path` is not a XLSX file.
            FileNotFoundError: If `path` is not in an existing directory.
        """
        xlsx_ext = ".xlsx"
        if not path.endswith(xlsx_ext):
            raise ValueError(f"Path '{path}' does not end with '{xlsx_ext}'")

        dir = os.path.dirname(path)
        if dir and not os.path.isdir(dir):
            raise FileNotFoundError(f"Directory '{dir}' does not exist")

    @abstractmethod
    def run(self) -> None:
        pass


class Update(Command):
    """The `update` command updates an existing xlbudget file.

    Attributes: Class Attributes
        name (str): The command's CLI name.
        aliases (List[str]): The command's CLI aliases.

    Attributes:
        input (Optional[str]): The path to the input file, otherwise paste in terminal.
        format (inputformat.InputFormat): The input file format.
        year (Optional[str]): The year all transactions were made, only relevant if
            the input format is 'BMO_CC_ADOBE'.
    """

    name: str = "update"
    aliases: List[str] = ["u"]

    @classmethod
    def configure_args(cls, subparsers: _SubParsersAction) -> None:
        """Configures the argument parser for the `update` command.

        Args:
            subparsers (_SubParsersAction): The command `subparsers`.
        """
        parser = _add_parser(
            subparsers,
            name=cls.name,
            aliases=cls.aliases,
            help="update an existing xlbudget file",
            cmd_cls=Update,
        )

        # required arguments
        parser.add_argument(
            "format",
            action=GetInputFormats,
            choices=GetInputFormats.input_formats.keys(),
            help="select an input format",
        )

        # optional arguments
        parser.add_argument("-i", "--input", help="path to the input file")
        parser.add_argument(
            "-y",
            "--year",
            help="year that all transactions were made, only relevant if input format "
            "is 'BMO_CC_ADOBE'",
        )

    def __init__(self, args: Namespace) -> None:
        super().__init__(args)

        self._check_input(args.input, args.format, args.year)
        self.input = args.input
        self.format = args.format
        self.year = args.year

        logger.debug(f"instance variables: {vars(self)}")

    @staticmethod
    def _check_input(
        input: Optional[str], input_format: Optional[InputFormat], year: Optional[str]
    ) -> None:
        """Check that `input` and `year` are valid.

        Args:
            input (Optional[str]): The input path.
            input_format (Optional[InputFormat]): The input format.
            year (Optional[str]): The year of all transactions.

        Raises:
            ValueError: If `input` is not None and the wrong file extension or DNE.
            ValueError: If `year` is None when `input_format` is 'BMO_CC_ADOBE'.
        """
        if input is None:
            return

        in_ext = (".csv", ".tsv", ".txt")
        if not input.endswith(in_ext):
            raise ValueError(f"Input '{input}' does not end with one of '{in_ext}'")

        if not os.path.isfile(input):
            raise ValueError(f"Input '{input}' is not an existing file")

        # get key from value: https://stackoverflow.com/a/13149770
        if input_format is not None:
            # validate year
            format = list(GetInputFormats.input_formats.keys())[
                list(GetInputFormats.input_formats.values()).index(input_format)
            ]
            if format == "BMO_CC_ADOBE" and year is None:
                raise ValueError(f"Must specify 'year' argument when {format=}")

            # validate input file type in more detail
            if input_format.seperator == "\t" and not input.endswith(".tsv"):
                raise ValueError(f"Input file should be TSV for {format=}")

    def run(self) -> None:
        logger.info(f"Parsing input {self.input}")
        df = parse_input(self.input, self.format, self.year)
        logger.debug(f"input file: {df.shape=}, df.dtypes=\n{df.dtypes}")
        logger.debug(f"df.head()=\n{df.head()}")

        if os.path.exists(self.path):
            logger.info(f"Loading xlbudget file {self.path}")
            wb = load_workbook(self.path)
        else:
            logger.warning(f"xlbudget file {self.path} does not exist, creating")
            wb = Workbook()
            ws = wb.active
            # ignore type mismatch of active worksheet
            wb.remove(ws)  # type: ignore[arg-type]

        logger.info("Updating xlbudget file")
        update_xlbudget(wb, df)

        if not self.trial:
            logger.info(f"Saving xlbudget file to {self.path}")
            wb.save(self.path)
        else:
            logger.info(f"Trial run: not saving xlbudget file to {self.path}")


def get_command_classes() -> List[Type[Command]]:
    """Gets all classes that implement the `Command` abstract class.

    Returns:
        A[n] `List[Type[Command]]` of all command classes.
    """
    command_module = sys.modules[__name__]
    return [getattr(command_module, c.__name__) for c in Command.__subclasses__()]


def _add_parser(
    subparsers: _SubParsersAction,
    name: str,
    aliases: List[str],
    help: str,
    cmd_cls: Type[Command],
) -> ArgumentParser:
    """Adds an argument parser for a command. Any configuration that is common
    across commands should go here.

    Args:
        subparsers (_SubParsersAction): The subparsers object.
        name (str): The command name.
        aliases (List[str]): The command aliases.
        help (str): The command help message.
        cmd_cls (Type[Command]): The command class.

    Returns:
        A[n] `ArgumentParser` for a command.
    """
    parser = subparsers.add_parser(name, aliases=aliases, help=help)

    # initialize the command with args.init(...)
    parser.set_defaults(init=cmd_cls)

    return parser
